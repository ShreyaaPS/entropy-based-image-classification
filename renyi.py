import cv2
import os
from numpy import array
import numpy as np
from openpyxl import Workbook
workbooks=Workbook()
sheet=workbooks.active
workbooks.save(filename='renyien.xlsx')
from openpyxl import load_workbook
wb=load_workbook('renyien.xlsx')
ws=wb.active
alpha=[2,3,4]
folder=r'C:\Users\SHREYA P S\OneDrive\Desktop\natural_images'
for j in alpha:
    for foldername in os.listdir(folder):
        images=[]
        file=os.path.join(folder,foldername)
        for filename in os.listdir(file):
            img=cv2.imread(os.path.join(file,filename))
            if img is not None:
                images.append(img)
        c=ws.max_column + 1        
        r=2
        for i in images:
            gray=cv2.cvtColor(i,cv2.COLOR_BGR2GRAY)
            ar=array(gray)
            arr=ar.flatten()
            value,counts=np.unique(arr,return_counts=True)
            prob=counts/arr.size
            alphaprob=prob ** j
            summation=alphaprob.sum()
            log=np.log2(summation)
            parameter=1/(1-j)
            entropy=parameter * log
            ws.cell(row=r,column=c).value=entropy
            r=r+1
            wb.save('renyien.xlsx')
